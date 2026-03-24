from openpyxl import Workbook, load_workbook

wb = load_workbook(filename = 'Data/Cartriges.xlsx')
sheet_ranges = wb['КАРТРИДЖИ']
ws = wb.active
cell_range = ws['A1':'A65']
for row in ws.iter_rows(min_row=6, max_col=1, max_row=65):
    for cell in row:
        print(cell.value)

