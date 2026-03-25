from openpyxl import Workbook, load_workbook

wb = load_workbook(filename = 'Data/Cartriges.xlsx')
sheet_ranges = wb['КАРТРИДЖИ']
ws = wb.active
cell_range = ws['A6':'A65']
names = list()
for row in cell_range:
    for cell in row:
        names.append(cell.value)

print(names)


